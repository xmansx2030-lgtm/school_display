from __future__ import annotations

from datetime import datetime, time
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from schedule.models import (
    ClassLesson,
    DaySchedule,
    Period,
    SchoolClass,
    SchoolSettings,
    Subject,
    Teacher,
)


DAY_NAMES = {
    "الاثنين": 1,
    "الثلاثاء": 2,
    "الأربعاء": 3,
    "الخميس": 4,
    "الجمعة": 5,
    "السبت": 6,
    "الأحد": 7,
    "1": 1,
    "2": 2,
    "3": 3,
    "4": 4,
    "5": 5,
    "6": 6,
    "7": 7,
}
DAY_LABELS = {value: key for key, value in DAY_NAMES.items() if not key.isdigit()}
MAX_IMPORT_ROWS = 5000
EXPECTED_HEADERS = {
    "المعلمون": ["اسم المعلم"],
    "الفصول": ["اسم الفصل"],
    "المواد": ["اسم المادة"],
    "الجدول": [
        "اليوم",
        "رقم الحصة",
        "وقت البداية",
        "وقت النهاية",
        "الفصل",
        "المادة",
        "المعلم",
    ],
}
TIMETABLE_COLUMNS = {
    "day": "اليوم",
    "period_index": "رقم الحصة",
    "starts_at": "وقت البداية",
    "ends_at": "وقت النهاية",
    "class_name": "الفصل",
    "subject_name": "المادة",
    "teacher_name": "المعلم",
}


def _style_sheet(ws):
    ws.sheet_view.rightToLeft = True
    fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_template_bytes() -> bytes:
    wb = Workbook()
    teachers = wb.active
    teachers.title = "المعلمون"
    teachers.append(["اسم المعلم"])
    teachers.append(["أحمد محمد"])
    classes = wb.create_sheet("الفصول")
    classes.append(["اسم الفصل"])
    classes.append(["الأول أ"])
    subjects = wb.create_sheet("المواد")
    subjects.append(["اسم المادة"])
    subjects.append(["الرياضيات"])
    timetable = wb.create_sheet("الجدول")
    timetable.append(
        ["اليوم", "رقم الحصة", "وقت البداية", "وقت النهاية", "الفصل", "المادة", "المعلم"]
    )
    timetable.append(["الأحد", 1, "07:00", "07:45", "الأول أ", "الرياضيات", "أحمد محمد"])
    instructions = wb.create_sheet("تعليمات")
    instructions.append(["تعليمات الاستيراد"])
    instructions.append(["لا تغيّر أسماء الأوراق أو عناوين الأعمدة."])
    instructions.append(["صيغة الوقت المقبولة 07:00 أو 07:00:00."])
    instructions.append(["يمكن تكرار اليوم ورقم الحصة لفصول مختلفة بشرط تطابق الوقت."])
    instructions.append(["راجع الأخطاء في شاشة المعاينة قبل تنفيذ الاستيراد."])
    for ws in wb.worksheets:
        _style_sheet(ws)
        for column in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 4, 42)
            ws.column_dimensions[column[0].column_letter].width = width
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _text(value) -> str:
    return str(value or "").strip()


def _normalized_header(value) -> str:
    return " ".join(_text(value).split())


def _issue(*, sheet: str, row: int | None, column: str, message: str, value="") -> dict:
    return {
        "sheet": sheet,
        "row": row,
        "column": column,
        "message": message,
        "value": _text(value),
    }


def _issue_text(issue: dict) -> str:
    location = f"ورقة «{issue['sheet']}»"
    if issue.get("row"):
        location += f"، الصف {issue['row']}"
    if issue.get("column"):
        location += f"، عمود «{issue['column']}»"
    return f"{location}: {issue['message']}"


def _validate_headers(wb) -> list[dict]:
    issues = []
    for sheet_name, expected in EXPECTED_HEADERS.items():
        if sheet_name not in wb.sheetnames:
            if sheet_name == "الجدول":
                issues.append(
                    _issue(
                        sheet=sheet_name,
                        row=1,
                        column="",
                        message="الورقة غير موجودة. استخدم القالب المعتمد.",
                    )
                )
            continue
        actual = [
            _normalized_header(cell.value)
            for cell in list(wb[sheet_name].iter_rows(min_row=1, max_row=1))[0][: len(expected)]
        ]
        for index, expected_name in enumerate(expected):
            actual_name = actual[index] if index < len(actual) else ""
            if actual_name != expected_name:
                issues.append(
                    _issue(
                        sheet=sheet_name,
                        row=1,
                        column=expected_name,
                        message=f"العنوان المتوقع «{expected_name}» وليس «{actual_name or 'فارغ'}».",
                        value=actual_name,
                    )
                )
    return issues


def _parse_time(value) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    raw = _text(value)
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def _names_from_sheet(wb, sheet_name: str) -> tuple[list[str], bool]:
    if sheet_name not in wb.sheetnames:
        return [], False
    result, seen = [], set()
    exceeded = False
    for row_number, row in enumerate(
        wb[sheet_name].iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if row_number > MAX_IMPORT_ROWS + 1:
            exceeded = True
            break
        name = _text(row[0] if row else "")
        if name and name.casefold() not in seen:
            result.append(name)
            seen.add(name.casefold())
    return result, exceeded


def parse_workbook(uploaded_file) -> dict:
    try:
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        issue = _issue(
            sheet="الملف",
            row=None,
            column="",
            message=f"تعذر قراءة ملف Excel: {exc}",
        )
        return {
            "rows": [],
            "errors": [_issue_text(issue)],
            "issues": [issue],
            "warnings": [],
            "entities": {},
            "summary": {},
        }

    issues = _validate_headers(wb)
    if "الجدول" not in wb.sheetnames:
        return {
            "rows": [],
            "errors": [_issue_text(issue) for issue in issues],
            "issues": issues,
            "warnings": [],
            "entities": {},
            "summary": {},
        }

    teachers, teachers_exceeded = _names_from_sheet(wb, "المعلمون")
    classes, classes_exceeded = _names_from_sheet(wb, "الفصول")
    subjects, subjects_exceeded = _names_from_sheet(wb, "المواد")
    entities = {"teachers": teachers, "classes": classes, "subjects": subjects}
    rows, seen_slots, slot_times, teacher_slots = [], set(), {}, {}
    for label, exceeded in (
        ("المعلمون", teachers_exceeded),
        ("الفصول", classes_exceeded),
        ("المواد", subjects_exceeded),
    ):
        if exceeded:
            issues.append(
                _issue(
                    sheet=label,
                    row=MAX_IMPORT_ROWS + 2,
                    column="",
                    message=f"تجاوزت الورقة الحد الأقصى ({MAX_IMPORT_ROWS} صف).",
                )
            )
    for row_number, values in enumerate(wb["الجدول"].iter_rows(min_row=2, values_only=True), start=2):
        if row_number > MAX_IMPORT_ROWS + 1:
            issues.append(
                _issue(
                    sheet="الجدول",
                    row=row_number,
                    column="",
                    message=f"تجاوز الملف الحد الأقصى ({MAX_IMPORT_ROWS} صف).",
                )
            )
            break
        padded = tuple(values or ()) + (None,) * max(0, 7 - len(tuple(values or ())))
        day_raw, index_raw, start_raw, end_raw, class_raw, subject_raw, teacher_raw = padded[:7]
        if not any(_text(value) for value in padded[:7]):
            continue
        row_issues = []
        day = DAY_NAMES.get(_text(day_raw))
        try:
            period_index = int(index_raw)
        except (TypeError, ValueError):
            period_index = 0
        starts_at, ends_at = _parse_time(start_raw), _parse_time(end_raw)
        class_name, subject_name, teacher_name = _text(class_raw), _text(subject_raw), _text(teacher_raw)
        if day is None:
            row_issues.append(("day", "اسم اليوم غير صحيح", day_raw))
        if period_index < 1 or period_index > 20:
            row_issues.append(("period_index", "رقم الحصة يجب أن يكون بين 1 و20", index_raw))
        if starts_at is None:
            row_issues.append(("starts_at", "وقت البداية غير صحيح", start_raw))
        if ends_at is None:
            row_issues.append(("ends_at", "وقت النهاية غير صحيح", end_raw))
        if starts_at and ends_at and starts_at >= ends_at:
            row_issues.append(("ends_at", "وقت النهاية يجب أن يكون بعد البداية", end_raw))
        if not class_name:
            row_issues.append(("class_name", "الفصل مطلوب", class_raw))
        if not subject_name:
            row_issues.append(("subject_name", "المادة مطلوبة", subject_raw))
        if not teacher_name:
            row_issues.append(("teacher_name", "المعلم مطلوب", teacher_raw))
        if day and period_index and starts_at and ends_at:
            slot = (day, period_index)
            previous = slot_times.get(slot)
            if previous and previous != (starts_at, ends_at):
                row_issues.append(
                    ("starts_at", "وقت الحصة لا يطابق الصفوف الأخرى لنفس اليوم ورقم الحصة", start_raw)
                )
            else:
                slot_times[slot] = (starts_at, ends_at)
            lesson_slot = (day, period_index, class_name.casefold())
            if lesson_slot in seen_slots:
                row_issues.append(("class_name", "الحصة مكررة لنفس الفصل", class_raw))
            seen_slots.add(lesson_slot)
            if teacher_name:
                teacher_slot = (day, period_index, teacher_name.casefold())
                previous_teacher = teacher_slots.get(teacher_slot)
                if previous_teacher and previous_teacher[0] != class_name.casefold():
                    row_issues.append(
                        (
                            "teacher_name",
                            f"المعلم مرتبط في الوقت نفسه بالفصل «{previous_teacher[1]}» في الصف {previous_teacher[2]}",
                            teacher_raw,
                        )
                    )
                else:
                    teacher_slots[teacher_slot] = (class_name.casefold(), class_name, row_number)

        normalized_issues = [
            _issue(
                sheet="الجدول",
                row=row_number,
                column=TIMETABLE_COLUMNS[field],
                message=message,
                value=value,
            )
            for field, message, value in row_issues
        ]
        issues.extend(normalized_issues)
        normalized = {
            "row": row_number,
            "day": day,
            "day_label": DAY_LABELS.get(day, _text(day_raw)),
            "period_index": period_index,
            "starts_at": starts_at.strftime("%H:%M") if starts_at else "",
            "ends_at": ends_at.strftime("%H:%M") if ends_at else "",
            "class_name": class_name,
            "subject_name": subject_name,
            "teacher_name": teacher_name,
            "errors": [issue["message"] for issue in normalized_issues],
            "issues": normalized_issues,
        }
        rows.append(normalized)
    if not rows:
        issues.append(
            _issue(
                sheet="الجدول",
                row=None,
                column="",
                message="ورقة الجدول لا تحتوي أي حصص.",
            )
        )
    errors = [_issue_text(issue) for issue in issues]
    return {
        "rows": rows,
        "errors": errors,
        "issues": issues,
        "warnings": [],
        "entities": entities,
        "summary": {
            "rows": len(rows),
            "valid_rows": sum(1 for row in rows if not row["errors"]),
            "invalid_rows": sum(1 for row in rows if row["errors"]),
            "teachers": len({row["teacher_name"].casefold() for row in rows if row["teacher_name"]}),
            "subjects": len({row["subject_name"].casefold() for row in rows if row["subject_name"]}),
            "classes": len({row["class_name"].casefold() for row in rows if row["class_name"]}),
        },
    }


def build_import_impact(*, school, parsed: dict) -> dict:
    """Describe what a valid import will create or update without writing data."""
    settings_obj = SchoolSettings.objects.filter(school=school).first()
    teacher_names = {row["teacher_name"] for row in parsed.get("rows", []) if row["teacher_name"]}
    subject_names = {row["subject_name"] for row in parsed.get("rows", []) if row["subject_name"]}
    class_names = {row["class_name"] for row in parsed.get("rows", []) if row["class_name"]}
    for name in (parsed.get("entities") or {}).get("teachers", []):
        teacher_names.add(name)
    for name in (parsed.get("entities") or {}).get("subjects", []):
        subject_names.add(name)
    for name in (parsed.get("entities") or {}).get("classes", []):
        class_names.add(name)

    existing_teachers = {
        name.casefold() for name in Teacher.objects.filter(school=school).values_list("name", flat=True)
    }
    existing_subjects = {
        name.casefold() for name in Subject.objects.filter(school=school).values_list("name", flat=True)
    }
    existing_classes = set()
    existing_slots = set()
    if settings_obj is not None:
        existing_classes = {
            name.casefold()
            for name in SchoolClass.objects.filter(settings=settings_obj).values_list("name", flat=True)
        }
        existing_slots = set(
            ClassLesson.objects.filter(settings=settings_obj).values_list(
                "weekday", "period_index", "school_class__name"
            )
        )
    imported_slots = {
        (row["day"], row["period_index"], row["class_name"])
        for row in parsed.get("rows", [])
        if not row.get("errors")
    }
    update_count = sum(1 for slot in imported_slots if slot in existing_slots)
    return {
        "new_teachers": sum(1 for name in teacher_names if name.casefold() not in existing_teachers),
        "existing_teachers": sum(1 for name in teacher_names if name.casefold() in existing_teachers),
        "new_subjects": sum(1 for name in subject_names if name.casefold() not in existing_subjects),
        "existing_subjects": sum(1 for name in subject_names if name.casefold() in existing_subjects),
        "new_classes": sum(1 for name in class_names if name.casefold() not in existing_classes),
        "existing_classes": sum(1 for name in class_names if name.casefold() in existing_classes),
        "new_lessons": max(0, len(imported_slots) - update_count),
        "updated_lessons": update_count,
    }


def _get_or_create_named(model, *, name: str, **scope):
    existing = model.objects.filter(**scope, name__iexact=name).first()
    if existing:
        return existing, False
    return model.objects.create(**scope, name=name), True


@transaction.atomic
def apply_import(*, school, parsed: dict) -> dict:
    if parsed.get("errors"):
        raise ValueError("لا يمكن استيراد ملف يحتوي أخطاء.")
    settings_obj, _ = SchoolSettings.objects.get_or_create(school=school, defaults={"name": school.name})
    entities = parsed.get("entities") or {}
    undo = {
        "created_teachers": [],
        "created_subjects": [],
        "created_classes": [],
        "created_days": [],
        "updated_days": [],
        "created_periods": [],
        "updated_periods": [],
        "created_lessons": [],
        "updated_lessons": [],
    }

    def ensure_teacher(name):
        obj, created = _get_or_create_named(Teacher, school=school, name=name)
        if created:
            undo["created_teachers"].append(obj.pk)
        return obj

    def ensure_subject(name):
        obj, created = _get_or_create_named(Subject, school=school, name=name)
        if created:
            undo["created_subjects"].append(obj.pk)
        return obj

    def ensure_class(name):
        obj, created = _get_or_create_named(SchoolClass, settings=settings_obj, name=name)
        if created:
            undo["created_classes"].append(obj.pk)
        return obj

    for name in entities.get("teachers", []):
        ensure_teacher(name)
    for name in entities.get("subjects", []):
        ensure_subject(name)
    for name in entities.get("classes", []):
        ensure_class(name)

    period_slots, imported_lessons = {}, 0
    for row in parsed["rows"]:
        teacher = ensure_teacher(row["teacher_name"])
        subject = ensure_subject(row["subject_name"])
        school_class = ensure_class(row["class_name"])
        day, day_created = DaySchedule.objects.get_or_create(
            settings=settings_obj,
            weekday=row["day"],
            defaults={"is_active": True, "periods_count": row["period_index"]},
        )
        if day_created:
            undo["created_days"].append(day.pk)
        if not day.is_active or day.periods_count < row["period_index"]:
            if not day_created and not any(item["id"] == day.pk for item in undo["updated_days"]):
                undo["updated_days"].append(
                    {"id": day.pk, "is_active": day.is_active, "periods_count": day.periods_count}
                )
            day.is_active = True
            day.periods_count = max(day.periods_count, row["period_index"])
            day.save(update_fields=("is_active", "periods_count"))
        slot = (row["day"], row["period_index"])
        if slot not in period_slots:
            period, period_created = Period.objects.get_or_create(
                day=day,
                index=row["period_index"],
                defaults={
                    "starts_at": datetime.strptime(row["starts_at"], "%H:%M").time(),
                    "ends_at": datetime.strptime(row["ends_at"], "%H:%M").time(),
                },
            )
            if period_created:
                undo["created_periods"].append(period.pk)
            else:
                starts_at = datetime.strptime(row["starts_at"], "%H:%M").time()
                ends_at = datetime.strptime(row["ends_at"], "%H:%M").time()
                if period.starts_at != starts_at or period.ends_at != ends_at:
                    undo["updated_periods"].append(
                        {
                            "id": period.pk,
                            "starts_at": period.starts_at.strftime("%H:%M:%S"),
                            "ends_at": period.ends_at.strftime("%H:%M:%S"),
                        }
                    )
                    period.starts_at = starts_at
                    period.ends_at = ends_at
                    period.save(update_fields=("starts_at", "ends_at"))
            period_slots[slot] = True
        lesson, lesson_created = ClassLesson.objects.get_or_create(
            settings=settings_obj,
            weekday=row["day"],
            period_index=row["period_index"],
            school_class=school_class,
            defaults={"subject": subject, "teacher": teacher, "is_active": True},
        )
        if lesson_created:
            undo["created_lessons"].append(lesson.pk)
        elif (
            lesson.subject_id != subject.pk
            or lesson.teacher_id != teacher.pk
            or not lesson.is_active
        ):
            undo["updated_lessons"].append(
                {
                    "id": lesson.pk,
                    "subject_id": lesson.subject_id,
                    "teacher_id": lesson.teacher_id,
                    "is_active": lesson.is_active,
                }
            )
            lesson.subject = subject
            lesson.teacher = teacher
            lesson.is_active = True
            lesson.save(update_fields=("subject", "teacher", "is_active"))
        imported_lessons += 1
    return {
        "lessons": imported_lessons,
        "teachers": Teacher.objects.filter(school=school).count(),
        "subjects": Subject.objects.filter(school=school).count(),
        "classes": SchoolClass.objects.filter(settings=settings_obj).count(),
        "undo": undo,
    }


@transaction.atomic
def undo_import(*, school, undo: dict) -> dict:
    """Reverse one import without touching records that existed before it."""
    settings_obj = SchoolSettings.objects.get(school=school)
    restored = 0

    for item in undo.get("updated_lessons", []):
        restored += ClassLesson.objects.filter(
            pk=item["id"], settings=settings_obj
        ).update(
            subject_id=item["subject_id"],
            teacher_id=item["teacher_id"],
            is_active=item["is_active"],
        )
    deleted_lessons, _ = ClassLesson.objects.filter(
        settings=settings_obj,
        pk__in=undo.get("created_lessons", []),
    ).delete()

    day_ids = list(DaySchedule.objects.filter(settings=settings_obj).values_list("pk", flat=True))
    for item in undo.get("updated_periods", []):
        restored += Period.objects.filter(pk=item["id"], day_id__in=day_ids).update(
            starts_at=datetime.strptime(item["starts_at"], "%H:%M:%S").time(),
            ends_at=datetime.strptime(item["ends_at"], "%H:%M:%S").time(),
        )
    deleted_periods, _ = Period.objects.filter(
        day_id__in=day_ids,
        pk__in=undo.get("created_periods", []),
    ).delete()

    for item in undo.get("updated_days", []):
        restored += DaySchedule.objects.filter(pk=item["id"], settings=settings_obj).update(
            is_active=item["is_active"],
            periods_count=item["periods_count"],
        )
    deleted_days, _ = DaySchedule.objects.filter(
        settings=settings_obj,
        pk__in=undo.get("created_days", []),
    ).delete()

    created_class_ids = undo.get("created_classes", [])
    created_subject_ids = undo.get("created_subjects", [])
    created_teacher_ids = undo.get("created_teachers", [])
    deleted_classes, _ = SchoolClass.objects.filter(
        settings=settings_obj,
        pk__in=created_class_ids,
        class_lessons__isnull=True,
        periods__isnull=True,
    ).delete()
    deleted_subjects, _ = Subject.objects.filter(
        school=school,
        pk__in=created_subject_ids,
        class_lessons__isnull=True,
        periods__isnull=True,
    ).delete()
    deleted_teachers, _ = Teacher.objects.filter(
        school=school,
        pk__in=created_teacher_ids,
        class_lessons__isnull=True,
        periods__isnull=True,
    ).delete()
    return {
        "restored": restored,
        "deleted": (
            deleted_lessons
            + deleted_periods
            + deleted_days
            + deleted_classes
            + deleted_subjects
            + deleted_teachers
        ),
    }
